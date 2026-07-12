"""Tests for lyceum.doc_writer — the assistant's document-writing engine
(model text -> real .docx/.xlsx files with live formulas)."""
import os
import tempfile
import unittest
from datetime import datetime

from lyceum import doc_writer as dw


class ParseTableTest(unittest.TestCase):
    def test_title_headers_rows(self):
        reply = "July Budget\nItem|Amount\nRent|900\nFood|250"
        title, headers, rows = dw.parse_table(reply)
        self.assertEqual(title, "July Budget")
        self.assertEqual(headers, ["Item", "Amount"])
        self.assertEqual(rows, [["Rent", 900.0], ["Food", 250.0]])

    def test_tolerates_currency_and_junk_lines(self):
        reply = ("Budget\n```\nItem|Amount\n---|---\nRent|$1,200\n"
                 "Food|250\n```")
        _t, headers, rows = dw.parse_table(reply)
        self.assertEqual(headers, ["Item", "Amount"])
        self.assertEqual(rows, [["Rent", 1200.0], ["Food", 250.0]])

    def test_empty_reply(self):
        self.assertEqual(dw.parse_table(""), (None, [], []))

    def test_refusal_detector(self):
        self.assertTrue(dw.looks_like_refusal("I cannot write that."))
        self.assertFalse(dw.looks_like_refusal("Dear Landlord,"))


@unittest.skipIf(dw._openpyxl is None, "openpyxl not installed")
class WriteXlsxTest(unittest.TestCase):
    def test_writes_rows_and_sum_formula(self):
        d = tempfile.mkdtemp()
        p = os.path.join(d, "b.xlsx")
        dw.write_table_xlsx(p, "Budget", ["Item", "Amount"],
                            [["Rent", 900.0], ["Food", 250.0]])
        import openpyxl
        ws = openpyxl.load_workbook(p).active
        self.assertEqual(ws.title, "Budget")
        self.assertEqual(ws.cell(row=2, column=1).value, "Rent")
        self.assertEqual(ws.cell(row=4, column=2).value, "=SUM(B2:B3)")
        self.assertEqual(ws.cell(row=4, column=1).value, "TOTAL")

    def test_sheet_title_with_forbidden_chars(self):
        # ':' is illegal in Excel tab names — must not abort the write.
        d = tempfile.mkdtemp()
        p = os.path.join(d, "c.xlsx")
        dw.write_table_xlsx(p, "Budget: July [draft]", ["Item", "Amount"],
                            [["Rent", 900.0]])
        import openpyxl
        ws = openpyxl.load_workbook(p).active
        self.assertEqual(ws.title, "Budget July draft")

    def test_text_only_table_gets_no_total(self):
        d = tempfile.mkdtemp()
        p = os.path.join(d, "t.xlsx")
        dw.write_table_xlsx(p, "Roles", ["Role", "Person"],
                            [["QA", "Shannon"], ["Dev", "Claude"]])
        import openpyxl
        ws = openpyxl.load_workbook(p).active
        self.assertEqual(ws.max_row, 3)   # header + 2 rows, no TOTAL


@unittest.skipIf(dw._docx is None, "python-docx not installed")
class WriteDocxTest(unittest.TestCase):
    def test_paragraph_blocks(self):
        d = tempfile.mkdtemp()
        p = os.path.join(d, "l.docx")
        dw.write_letter_docx(p, None, "Dear Sir,\n\nFirst.\n\nSecond.")
        import docx
        texts = [pp.text for pp in docx.Document(p).paragraphs if pp.text]
        self.assertEqual(texts, ["Dear Sir,", "First.", "Second."])


class FilenameTest(unittest.TestCase):
    def test_sanitizes_and_dates(self):
        when = datetime(2026, 7, 4, 14, 32)
        name = dw.suggest_filename("docx", 'rent: "receipt"/letter?',
                                   when=when)
        self.assertEqual(name, "rent receipt letter 2026-07-04 1432.docx")

    def test_defaults(self):
        when = datetime(2026, 7, 4, 9, 5)
        self.assertEqual(dw.suggest_filename("xlsx", None, when=when),
                         "Spreadsheet 2026-07-04 0905.xlsx")


class ComputeTotalsTest(unittest.TestCase):
    """compute_totals() evaluates a table's column SUMs via the formula
    engine (no Excel needed) — the doc_writer ↔ formula.py wiring."""

    def test_sums_numeric_columns(self):
        headers = ["Item", "Amount"]
        rows = [["Rent", 900.0], ["Food", 250.0], ["Gas", 120.0]]
        totals = dw.compute_totals(headers, rows)
        self.assertEqual(totals, {"Amount": 1270.0})

    def test_multiple_numeric_columns(self):
        headers = ["Task", "Hours", "Rate"]
        rows = [["A", 2.0, 20.0], ["B", 3.0, 25.0]]
        totals = dw.compute_totals(headers, rows)
        self.assertEqual(totals["Hours"], 5.0)
        self.assertEqual(totals["Rate"], 45.0)

    def test_text_only_table_has_no_totals(self):
        totals = dw.compute_totals(["Role", "Name"],
                                   [["QA", "Shannon"], ["Dev", "Claude"]])
        self.assertEqual(totals, {})

    def test_empty_rows(self):
        self.assertEqual(dw.compute_totals(["A"], []), {})


if __name__ == "__main__":
    unittest.main()
